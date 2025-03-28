
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def save_folder_structure_to_excel(subfolders, output_file='drive_folder_structure.xlsx'):
    """
    Save the folder structure to an Excel file with a proper hierarchical view.
    
    Args:
        subfolders: List of dictionaries containing folder details (from get_all_subfolders_multithreaded)
        output_file: Path to the output Excel file
    """
    print(f"Saving folder structure to {output_file}...")
    
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    
    # Create hierarchical tree view sheet
    tree_sheet = workbook.active
    tree_sheet.title = "Folder Tree"
    
    # Create flat list sheet with folder paths
    flat_sheet = workbook.create_sheet("Folder List")
    flat_sheet.append(["Folder ID", "Folder Name", "Full Path"])
    
    # Set column widths for flat sheet
    flat_sheet.column_dimensions['A'].width = 40
    flat_sheet.column_dimensions['B'].width = 40
    flat_sheet.column_dimensions['C'].width = 100
    
    # Add header formatting for flat sheet
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in flat_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add root folder to the tree
    tree_sheet.append(["Root Folder"])
    
    # Auto-adjust the first column width
    tree_sheet.column_dimensions['A'].width = 30
    
    # Configure reasonable widths for the hierarchy columns (up to 10 levels deep)
    for i in range(2, 12):  # Columns B through K (up to 10 levels of subfolders)
        col_letter = get_column_letter(i)
        tree_sheet.column_dimensions[col_letter].width = 30
    
    # Sort subfolders by path for better organization
    sorted_subfolders = sorted(subfolders, key=lambda x: x['path'])
    
    # Process each subfolder for the tree view
    for folder in sorted_subfolders:
        # Calculate the depth level
        path_parts = folder['path'].split('/')
        depth = len(path_parts)
        
        # Create a new row
        row = [""] * (depth + 1)  # +1 to account for 0-indexing
        row[depth] = folder['name']  # Place the folder name in the appropriate column
        
        # Add to tree view
        tree_sheet.append(row)
        
        # Add to flat list
        flat_sheet.append([folder['id'], folder['name'], folder['path']])
    
    # Freeze the header row in both sheets
    tree_sheet.freeze_panes = "A2"
    flat_sheet.freeze_panes = "A2"
    
    # Save the workbook
    workbook.save(output_file)
    print(f"Saved folder structure to {output_file}")
    
    # Also create a pandas dataframe version (useful for further processing)
    df = pd.DataFrame(sorted_subfolders)
    return df
