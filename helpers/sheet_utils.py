import io
import csv
import traceback

def extract_complete_sheet_text(file_data, file_name=None, file_url=None):
    """
    Converts a complete spreadsheet to text format optimized for AI processing.
    Extracts ALL rows, not just a sample, and formats the data in a way that
    preserves relationships and context.
    
    Args:
        file_data: BytesIO object containing the spreadsheet file
        file_name: Optional name of the file for context
        file_url: Optional URL of the file for context
        
    Returns:
        str: Complete text representation of the spreadsheet
    """
    try:
        # Try to use pandas if available (most robust method)
        try:
            import pandas as pd
            return extract_complete_sheet_with_pandas(file_data, file_name, file_url)
        except ImportError:
            # If pandas isn't available, use fallback methods
            pass
        
        # Create file-like object from bytes
        file_obj = io.BytesIO(file_data)
        
        # First try processing as CSV (simplest format)
        if file_name and file_name.lower().endswith('.csv'):
            return process_complete_csv(file_obj, file_name, file_url)
        
        # If not CSV, see if we can process Excel file
        try:
            # Try with openpyxl for modern Excel files
            return process_complete_xlsx_with_openpyxl(file_obj, file_name, file_url)
        except ImportError:
            try:
                # Try with xlrd for older Excel files
                return process_complete_xls_with_xlrd(file_obj, file_name, file_url)
            except ImportError:
                # If no Excel libraries are available, return helpful error
                return ("Unable to process Excel file: required libraries not installed. "
                        "Please install with: pip install pandas openpyxl xlrd")
    
    except Exception as e:
        error_details = traceback.format_exc()
        return (f"Error processing spreadsheet: {str(e)}\n\n"
                f"Try installing required dependencies with:\n"
                f"pip install pandas openpyxl xlrd\n\n"
                f"Error details:\n{error_details}")


def extract_complete_sheet_with_pandas(file_data, file_name=None, file_url=None):
    """Process complete spreadsheet with pandas if available"""
    import pandas as pd
    import numpy as np
    
    # Create file-like object
    file_obj = io.BytesIO(file_data)
    
    # Determine file type and read with appropriate method
    if file_name and file_name.lower().endswith('.csv'):
        df = pd.read_csv(file_obj)
    else:
        # Try with openpyxl first
        try:
            df = pd.read_excel(file_obj, engine='openpyxl')
        except:
            # If that fails, retry without specifying engine
            file_obj.seek(0)
            try:
                df = pd.read_excel(file_obj)
            except:
                # Last attempt with xlrd for older formats
                file_obj.seek(0)
                df = pd.read_excel(file_obj, engine='xlrd')
    
    # Start building output
    output_parts = []
    
    # Add metadata section
    metadata = []
    if file_name:
        metadata.append(f"Spreadsheet: {file_name}")
    if file_url:
        metadata.append(f"URL: {file_url}")
    
    # Add basic dataset information
    metadata.append(f"Rows: {len(df)}")
    metadata.append(f"Columns: {len(df.columns)} ({', '.join(df.columns)})")
    
    # Calculate some basic stats for context
    num_cols = df.select_dtypes(include=['number']).columns
    if len(num_cols) > 0:
        numeric_stats = []
        for col in num_cols:
            try:
                mean = df[col].mean()
                min_val = df[col].min()
                max_val = df[col].max()
                numeric_stats.append(f"{col} (range: {min_val} to {max_val}, avg: {mean:.2f})")
            except:
                pass
        
        if numeric_stats:
            metadata.append(f"Numeric columns: {'; '.join(numeric_stats)}")
    
    # Add metadata section to output
    output_parts.append("## METADATA ##")
    output_parts.append("\n".join(metadata))
    output_parts.append("## END METADATA ##")
    
    # Convert the entire DataFrame to a tabular text format
    output_parts.append("## DATA ##")
    
    # First add the column headers
    headers = df.columns.tolist()
    output_parts.append("|".join(str(header) for header in headers))
    
    # Then add all rows
    for _, row in df.iterrows():
        # Format each value and join with pipe separator
        formatted_values = []
        for col in headers:
            val = row[col]
            if pd.isna(val):
                formatted_val = ""  # Empty string for missing values
            elif isinstance(val, float):
                # Format floats to avoid scientific notation and limit decimal places
                formatted_val = f"{val:.6f}".rstrip('0').rstrip('.') if val == val else ""
            else:
                # For other types, convert to string and escape any pipe characters
                formatted_val = str(val).replace("|", "\\|")
            
            formatted_values.append(formatted_val)
        
        output_parts.append("|".join(formatted_values))
    
    output_parts.append("## END DATA ##")
    
    # Add narrative description for better AI understanding
    output_parts.append("## DESCRIPTION ##")
    
    description = []
    description.append(f"This spreadsheet contains {len(df)} records with {len(df.columns)} attributes per record.")
    
    # Add information about data types
    cat_cols = df.select_dtypes(include=['object', 'category']).columns
    date_cols = df.select_dtypes(include=['datetime']).columns
    
    if len(cat_cols) > 0:
        description.append(f"Categorical columns: {', '.join(cat_cols)}")
    
    if len(num_cols) > 0:
        description.append(f"Numerical columns: {', '.join(num_cols)}")
    
    if len(date_cols) > 0:
        description.append(f"Date columns: {', '.join(date_cols)}")
    
    # Try to identify potential relationships
    if len(df) > 5 and len(num_cols) >= 2:
        try:
            # Look for correlations between numerical columns
            corr_matrix = df[num_cols].corr()
            strong_correlations = []
            
            for i, col1 in enumerate(corr_matrix.columns):
                for j, col2 in enumerate(corr_matrix.columns):
                    if i < j:
                        corr_value = corr_matrix.iloc[i, j]
                        if abs(corr_value) > 0.7:
                            relation = "positively" if corr_value > 0 else "negatively"
                            strong_correlations.append(f"{col1} and {col2} are {relation} correlated")
            
            if strong_correlations:
                description.append("Potential relationships: " + "; ".join(strong_correlations))
        except:
            pass  # Skip correlation analysis if it fails
    
    # Add description to output
    output_parts.append("\n".join(description))
    output_parts.append("## END DESCRIPTION ##")
    
    # Return complete text
    return "\n\n".join(output_parts)


def process_complete_csv(file_obj, file_name=None, file_url=None):
    """Process a complete CSV file with basic csv module"""
    file_obj.seek(0)
    
    output_parts = []
    
    # Add metadata section
    metadata = []
    metadata.append("Format: CSV")
    
    # Try to decode as UTF-8
    try:
        content = file_obj.read().decode('utf-8')
    except UnicodeDecodeError:
        # If UTF-8 fails, try latin-1
        file_obj.seek(0)
        content = file_obj.read().decode('latin-1')
    
    reader = csv.reader(content.splitlines())
    rows = list(reader)
    
    if not rows:
        return "CSV file appears to be empty."
    
    # Get headers and data
    headers = rows[0]
    data_rows = rows[1:]
    
    # Update metadata
    metadata.append(f"Rows: {len(data_rows)}")
    metadata.append(f"Columns: {len(headers)} ({', '.join(headers)})")
    
    # Add metadata to output
    output_parts.append("## METADATA ##")
    output_parts.append("\n".join(metadata))
    output_parts.append("## END METADATA ##")
    
    # Add complete data
    output_parts.append("## DATA ##")
    
    # First add column headers
    output_parts.append("|".join(headers))
    
    # Then add all rows
    for row in data_rows:
        # Ensure row has enough values for all columns (pad if necessary)
        padded_row = row + [''] * (len(headers) - len(row))
        # Escape any pipe characters in the values
        escaped_row = [val.replace('|', '\\|') for val in padded_row[:len(headers)]]
        output_parts.append("|".join(escaped_row))
    
    output_parts.append("## END DATA ##")
    
    # Add basic description
    output_parts.append("## DESCRIPTION ##")
    description = [
        f"This CSV file contains {len(data_rows)} records with {len(headers)} columns.",
        f"The columns are: {', '.join(headers)}."
    ]
    output_parts.append("\n".join(description))
    output_parts.append("## END DESCRIPTION ##")
    
    return "\n\n".join(output_parts)


def process_complete_xlsx_with_openpyxl(file_obj, file_name=None, file_url=None):
    """Process complete Excel XLSX file with openpyxl"""
    import openpyxl
    
    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    
    output_parts = []
    
    # Add metadata section
    metadata = []
    if file_name:
        metadata.append(f"Spreadsheet: {file_name}")

    if file_url:
        metadata.append(f"URL: {file_url}")

    metadata.append("Format: Excel (XLSX)")
    metadata.append(f"Sheets: {', '.join(wb.sheetnames)}")
    
    # Get the first sheet
    sheet = wb.active
    
    # Get dimensions
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Get headers (first row)
    headers = []
    for col in range(1, max_col + 1):
        cell_value = sheet.cell(row=1, column=col).value
        headers.append(str(cell_value) if cell_value is not None else f"Column_{col}")
    
    # Update metadata
    metadata.append(f"Active Sheet: {sheet.title}")
    metadata.append(f"Rows: {max_row - 1}")
    metadata.append(f"Columns: {max_col} ({', '.join(headers)})")
    
    # Add metadata to output
    output_parts.append("## METADATA ##")
    output_parts.append("\n".join(metadata))
    output_parts.append("## END METADATA ##")
    
    # Add complete data
    output_parts.append("## DATA ##")
    
    # First add column headers
    output_parts.append("|".join(headers))
    
    # Then add all data rows
    for row in range(2, max_row + 1):  # Start from row 2 (after header)
        row_values = []
        for col in range(1, max_col + 1):
            value = sheet.cell(row=row, column=col).value
            # Format the cell value
            if value is None:
                formatted_val = ""
            elif isinstance(value, float):
                # Format floats to avoid scientific notation
                formatted_val = f"{value:.6f}".rstrip('0').rstrip('.')
            else:
                # Escape any pipe characters
                formatted_val = str(value).replace("|", "\\|")
            
            row_values.append(formatted_val)
        
        output_parts.append("|".join(row_values))
    
    output_parts.append("## END DATA ##")
    
    # Add basic description
    output_parts.append("## DESCRIPTION ##")
    description = [
        f"This Excel file contains {max_row - 1} records with {max_col} columns in sheet '{sheet.title}'.",
        f"The columns are: {', '.join(headers)}."
    ]
    output_parts.append("\n".join(description))
    output_parts.append("## END DESCRIPTION ##")
    
    wb.close()
    return "\n\n".join(output_parts)


def process_complete_xls_with_xlrd(file_obj, file_name=None, file_url=None):
    """Process complete old-format Excel XLS file with xlrd"""
    import xlrd
    
    file_obj.seek(0)
    wb = xlrd.open_workbook(file_contents=file_obj.read())
    
    output_parts = []
    
    # Add metadata section
    metadata = []

    metadata.append("Format: Excel (XLS)")
    metadata.append(f"Sheets: {', '.join(wb.sheet_names())}")
    
    # Get the first sheet
    sheet = wb.sheet_by_index(0)
    
    # Get dimensions
    num_rows = sheet.nrows
    num_cols = sheet.ncols
    
    if num_rows == 0:
        return "Excel file appears to be empty."
    
    # Get headers (first row)
    headers = []
    if num_rows > 0:
        for col in range(num_cols):
            cell_value = sheet.cell_value(0, col)
            headers.append(str(cell_value) if cell_value else f"Column_{col+1}")
    
    # Update metadata
    metadata.append(f"Active Sheet: {sheet.name}")
    metadata.append(f"Rows: {num_rows - 1}")
    metadata.append(f"Columns: {num_cols} ({', '.join(headers)})")
    
    # Add metadata to output
    output_parts.append("## METADATA ##")
    output_parts.append("\n".join(metadata))
    output_parts.append("## END METADATA ##")
    
    # Add complete data
    output_parts.append("## DATA ##")
    
    # First add column headers
    output_parts.append("|".join(headers))
    
    # Then add all data rows
    for row in range(1, num_rows):  # Start from row 1 (after header)
        row_values = []
        for col in range(num_cols):
            value = sheet.cell_value(row, col)
            # Format the cell value
            if value == '':
                formatted_val = ""
            elif isinstance(value, float):
                # Format floats to avoid scientific notation
                formatted_val = f"{value:.6f}".rstrip('0').rstrip('.')
            else:
                # Escape any pipe characters
                formatted_val = str(value).replace("|", "\\|")
            
            row_values.append(formatted_val)
        
        output_parts.append("|".join(row_values))
    
    output_parts.append("## END DATA ##")
    
    # Add basic description
    output_parts.append("## DESCRIPTION ##")
    description = [
        f"This Excel file contains {num_rows - 1} records with {num_cols} columns in sheet '{sheet.name}'.",
        f"The columns are: {', '.join(headers)}."
    ]
    output_parts.append("\n".join(description))
    output_parts.append("## END DESCRIPTION ##")
    
    return "\n\n".join(output_parts)